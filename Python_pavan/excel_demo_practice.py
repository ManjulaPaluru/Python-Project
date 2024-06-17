import openpyxl
Dict ={}
workbook = openpyxl.load_workbook("/Users/ssaguturu/Downloads/Python_excel_Demo.xlsx")
sheet = workbook.active
for i in range(1, sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(Dict)
