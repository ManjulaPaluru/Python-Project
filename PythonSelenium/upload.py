import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import wait, expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

def update_excel_data(file_path, search_term,column_name,new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == column_name:
            Dict["col"] = i
    for i in range(1, sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == search_term:
                Dict["row"] = i
    sheet.cell(row= Dict["row"],column= Dict["col"] ).value = new_value
    book.save(file_path)


file_path = "/Users/ssaguturu/Downloads/download.xlsx"
fruit_name = "Kivi"
new_value = "1000"
driver = webdriver.Chrome()
driver.implicitly_wait(20)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
# download a file
driver.find_element(By.ID,"downloadButton").click()

# edit the excel with updated value
update_excel_data(file_path, fruit_name, "price", new_value)

# upload the file
#if its windows based popup to upload file, first need to check in dom whether that button have type=file attribute ,
# if its not there check with dev to provide that.based on file_input send the file path to upload file
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver,5)
toast_locator = (By.XPATH,"//div[text()='Updated Excel Data Successfully.']")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text) # * symbol is useful for unpackingand refering that locator
priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actual_price)

assert actual_price == new_value